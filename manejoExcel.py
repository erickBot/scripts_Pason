import openpyxl
from datetime import datetime, date, time, timedelta
import calendar

list_vigencia = []
count_sensor = []

def run():

    name_file = "Lista_de_Sensores_HGAS.xlsx"
    file = openpyxl.load_workbook(name_file)
    #<file.sheetnames> devuelve los nombres de las hojas en una lista
    list_sheetbyName = file.sheetnames

    for name in list_sheetbyName:
        sheet = file.get_sheet_by_name(name)
        num_row = sheet.max_row
        celdas= sheet['H3':'H' + str(num_row)]
        dateNow = datetime.now()
        if (name == 'LEL'):
            vigencia = 730
        if (name == 'H2S'):
            vigencia = 365

        lookCell(celdas, dateNow, vigencia)

        #salta a las funciones para llenar las celdas con datos nuevos
        llenar_cell(sheet, list_vigencia, count_sensor, 'I3', 'I' + str(num_row))
        #guarda archivo
        list_vigencia.clear()
        count_sensor.clear()
    
    file.save(name_file)


def lookCell(celdas, dateNow, vigencia):
    expirados = 0
    porVencer = 0
    vigentes = 0

    for row in celdas:
        for cell in row:
            dateLast= cell.value
            dias = dateNow - dateLast
            diferencia = vigencia - dias.days 
            list_vigencia.append(diferencia)
    
            if diferencia < 0:
                expirados += 1
            if diferencia > 0 and diferencia < 30:
                porVencer += 1
            if diferencia > 30:
                vigentes += 1

    count_sensor.append(porVencer)
    count_sensor.append(expirados)
    count_sensor.append(vigentes)

    print(count_sensor)


def llenar_cell(sheet, list_vigencia, count_sensor, cellInit, cellEnd):
    i = 0
    cell_list = sheet[cellInit : cellEnd]
    for row in cell_list:
        for cell in row:
            cell.value = list_vigencia[i]
            i += 1

    sheet['R3'] = count_sensor[0] #porVencer
    sheet['R4'] = count_sensor[1] #expirados
    sheet['R5'] = count_sensor[2] #vigentes

if __name__ == "__main__":
    run()
